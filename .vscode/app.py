import tkinter
from tkinter import *
from tkinter import messagebox
import os
from openpyxl import Workbook
import openpyxl, xlrd
from openpyxl import load_workbook
import pathlib

# Creating a root for GUI
menu = Tk()
menu.geometry("580x375")
menu.title('B.A.T.E Internet')
# User databses
file = pathlib.Path('Users.xlsx')
if file.exists():
    pass
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"
    ws['A1'] = "User No."
    ws['B1'] = "Username"
    ws['C1'] = "Password"
    ws['D1'] = "Balance"
    ws['E1'] = "Mobile plans"
    ws['F1'] = "Domain name"
    ws['G1'] = "VPS packages"
    ws['H1'] = "VPN packages"

    wb.save('Users.xlsx')


# Creating and displaying a "Sign out" button
button = Button(menu, text="Sign out")
button.pack()

menu.mainloop()